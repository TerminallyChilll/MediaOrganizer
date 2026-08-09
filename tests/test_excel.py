from pathlib import Path

import openpyxl
import pandas as pd

from mediaorg import excel, scan
from mediaorg.excel import append_changes, plan_renames, read_library, write_library
from mediaorg.plan import NamingScheme
from mediaorg.scan import scan_movies, scan_tv


def make_movie(root, folder, files):
    d = root / folder
    d.mkdir(parents=True)
    for f in files:
        (d / f).write_text("x")


def test_scan_write_read_roundtrip(tmp_path):
    movies = tmp_path / "Movies"
    make_movie(movies, "The.Matrix.1999.1080p.BluRay.x264-RARBG",
               ["The.Matrix.1999.1080p.BluRay.x264-RARBG.mkv"])
    tv = tmp_path / "TV"
    show = tv / "The.Office.US.S01-S09.COMPLETE"
    (show / "Season 1").mkdir(parents=True)
    (show / "Season 1" / "The.Office.S01E01.720p.mkv").write_text("x")

    movie_rows = scan_movies(movies)
    tv_rows = scan_tv(tv)
    assert movie_rows[0]['Title'] == "The Matrix"
    assert movie_rows[0]['Year'] == 1999
    assert tv_rows[0]['Season'] == 1 and tv_rows[0]['Episode'] == 1
    assert tv_rows[0]['Episode File'] == "Season 1/The.Office.S01E01.720p.mkv"

    xlsx = tmp_path / "lib.xlsx"
    write_library(xlsx, movie_rows, tv_rows, movies, tv)
    df_movies, df_tv, meta = read_library(xlsx)
    assert df_movies is not None and len(df_movies) == 1
    assert df_tv is not None and len(df_tv) == 1
    assert meta['Movies Path'] == str(movies)
    assert 'Last Scan' in meta


def test_scan_is_read_only(tmp_path):
    movies = tmp_path / "Movies"
    movies.mkdir()
    (movies / "Loose.Movie.2020.mkv").write_text("x")  # loose file at root
    before = sorted(str(p) for p in tmp_path.rglob("*"))
    scan_movies(movies)
    assert sorted(str(p) for p in tmp_path.rglob("*")) == before


def test_fixed_column_overrides_flow_into_plan(tmp_path):
    movies = tmp_path / "Movies"
    make_movie(movies, "Teh.Matirx.1999.1080p", ["Teh.Matirx.1999.1080p.mkv"])
    xlsx = tmp_path / "lib.xlsx"
    write_library(xlsx, scan_movies(movies), [], movies, None)

    # User fixes the garbled title in Excel.
    wb = openpyxl.load_workbook(xlsx)
    ws = wb['Movies']
    headers = [c.value for c in ws[1]]
    ws.cell(row=2, column=headers.index('Title Fixed') + 1, value="The Matrix")
    wb.save(xlsx)

    df_movies, df_tv, _ = read_library(xlsx)
    plan = plan_renames(df_movies, movies, df_tv, None, NamingScheme())
    dsts = sorted(o.dst.relative_to(movies).as_posix() for o in plan.ops)
    assert dsts == [
        "Teh.Matirx.1999.1080p/The Matrix (1999) [1080p].mkv",
        "The Matrix (1999) [1080p]",
    ]
    # Children before parents: file rename precedes folder rename.
    assert plan.ops[0].src.name == "Teh.Matirx.1999.1080p.mkv"
    assert plan.ops[-1].src.name == "Teh.Matirx.1999.1080p"


def test_plan_renames_idempotent_when_clean(tmp_path):
    movies = tmp_path / "Movies"
    make_movie(movies, "The Matrix (1999) [1080p]",
               ["The Matrix (1999) [1080p].mkv"])
    xlsx = tmp_path / "lib.xlsx"
    write_library(xlsx, scan_movies(movies), [], movies, None)
    df_movies, _, _ = read_library(xlsx)
    plan = plan_renames(df_movies, movies, None, None, NamingScheme())
    assert plan.ops == []


def test_tv_rename_children_first_and_season_year(tmp_path):
    tv = tmp_path / "TV"
    show = tv / "The.Office.US.1080p.WEB"
    (show / "Season 1").mkdir(parents=True)
    (show / "Season 1" / "The.Office.S01E01.Pilot.720p.mkv").write_text("x")
    xlsx = tmp_path / "lib.xlsx"
    write_library(xlsx, [], scan_tv(tv), None, tv)

    _, df_tv, _ = read_library(xlsx)
    scheme = NamingScheme()
    scheme.tv_season_include_year = False
    plan = plan_renames(None, None, df_tv, tv, scheme)
    kinds = [(o.src.name, o.dst.name) for o in plan.ops]
    # Episode renamed in place, then show folder renamed; episode name keeps
    # show title + code + episode title.
    assert kinds[0][1] == "The Office (US) S01E01 - Pilot [720p].mkv"
    assert kinds[-1] == ("The.Office.US.1080p.WEB", "The Office (US)")
    for op in plan.ops:
        assert op.dst.parent == op.src.parent  # pure renames only


def test_companions_renamed_with_their_episode(tmp_path):
    tv = tmp_path / "TV"
    show = tv / "Show"
    (show / "Season 1").mkdir(parents=True)
    (show / "Season 1" / "Show.S01E01.720p.mkv").write_text("x")
    (show / "Season 1" / "Show.S01E01.en.srt").write_text("x")  # code match
    (show / "Season 1" / "Show.S01E02.mkv").write_text("x")
    xlsx = tmp_path / "lib.xlsx"
    write_library(xlsx, [], scan_tv(tv), None, tv)
    _, df_tv, _ = read_library(xlsx)
    plan = plan_renames(None, None, df_tv, tv, NamingScheme())
    names = {o.src.name: o.dst.name for o in plan.ops}
    assert names["Show.S01E01.en.srt"] == "Show S01E01 [720p].en.srt"


def test_a_numbered_sibling_is_not_a_companion(tmp_path):
    """"Film10.en" starts with "Film1" but belongs to a different film.

    Movies have no SxxEyy code to fall back on, so the prefix test is the only
    thing standing between Film1's rename and Film10's subtitle.
    """
    root = tmp_path / "Movies"
    root.mkdir()
    for name in ("Film1.mkv", "Film1.en.srt", "Film10.en.srt"):
        (root / name).write_text(name)

    ops = excel._companion_ops(root / "Film1.mkv", "Film1.mkv", "Pilot.mkv")

    assert {o.src.name: o.dst.name for o in ops} == {
        "Film1.en.srt": "Pilot.en.srt"}


def test_show_root_scan_does_not_rename_season_folders_as_shows(tmp_path):
    # Scanner pointed at a single show's root: "Show Folder" == "Season 1".
    root = tmp_path / "The Office"
    (root / "Season 1").mkdir(parents=True)
    (root / "Season 1" / "The.Office.S01E01.mkv").write_text("x")
    xlsx = tmp_path / "lib.xlsx"
    write_library(xlsx, [], scan_tv(root), None, root)
    _, df_tv, _ = read_library(xlsx)
    plan = plan_renames(None, None, df_tv, root, NamingScheme())
    # No op may rename the "Season 1" folder into a show-style name.
    for op in plan.ops:
        assert op.src.name != "Season 1"


def test_append_mode_dedupes(tmp_path):
    movies = tmp_path / "Movies"
    make_movie(movies, "Movie.A.2020", ["Movie.A.2020.mkv"])
    xlsx = tmp_path / "lib.xlsx"
    write_library(xlsx, scan_movies(movies), [], movies, None)
    make_movie(movies, "Movie.B.2021", ["Movie.B.2021.mkv"])
    write_library(xlsx, scan_movies(movies), [], movies, None, append=True)
    df_movies, _, _ = read_library(xlsx)
    assert sorted(df_movies['Folder Name']) == ["Movie.A.2020", "Movie.B.2021"]
    assert len(df_movies) == 2


def test_append_changes_sheet(tmp_path):
    xlsx = tmp_path / "lib.xlsx"
    write_library(xlsx, [{'Folder Name': 'X', 'Title': 'X'}], [], None, None)
    entries = [{'op': 'move', 'src': '/a/old.mkv', 'dst': '/a/new.mkv', 'ts': 1700000000.0}]
    append_changes(xlsx, entries)
    append_changes(xlsx, entries)
    df = pd.read_excel(xlsx, sheet_name='Changes')
    assert len(df) == 2
    assert df.iloc[0]['From'] == '/a/old.mkv'
    # Other sheets survived the append.
    assert read_library(xlsx)[0] is not None


def test_root_show_rows_use_selected_folder_name(tmp_path):
    # Recursive TV scans record root-level episodes with Show Folder '.'
    # — the rename planner must use the selected folder's name as the
    # title and never rename the root itself (PR review regression).
    import pandas as pd
    root = tmp_path / "The Office"
    root.mkdir()
    (root / "The.Office.S01E01.720p.mkv").write_text("x")
    df_tv = pd.DataFrame([{
        'Show Folder': '.', 'Folder Fixed': '',
        'Title': 'The Office', 'Title Fixed': '',
        'Season': 1, 'Season Year': '', 'Episode': 1,
        'Episode File': 'The.Office.S01E01.720p.mkv', 'File Fixed': '',
        'Quality': '720p', 'Quality Fixed': '', 'Size (GB)': 0.0,
    }])
    plan = plan_renames(None, None, df_tv, root, NamingScheme())
    names = {o.src.name: o.dst.name for o in plan.ops}
    assert names["The.Office.S01E01.720p.mkv"] == "The Office S01E01 [720p].mkv"
    # The selected root folder itself is never a rename target.
    assert all(o.src != root for o in plan.ops)


# --- Routing / naming regressions -------------------------------------------

import unicodedata


def test_unclassifiable_files_are_not_renamed_to_s00e00(tmp_path):
    """Extras and misplaced movies used to become 'Show S00E00.mkv'."""
    show = tmp_path / "Show"
    (show / "Extras").mkdir(parents=True)
    (show / "Season 1").mkdir(parents=True)
    (show / "Season 1" / "Show.S01E01.mkv").write_text("x")
    (show / "Extras" / "Behind The Scenes.mkv").write_text("x")

    rows = scan.scan_recursive_tv(tmp_path)
    unclassified = [r for r in rows if r['Season'] == '']
    assert unclassified, "expected the extra to be recorded as unclassified"

    df_tv = pd.DataFrame(rows)
    plan = excel.plan_renames(None, None, df_tv, tmp_path, NamingScheme())
    touched = {str(o.src) for o in plan.ops}
    assert not any("Behind The Scenes" in t for t in touched)


def test_appledouble_does_not_block_the_real_rename(tmp_path):
    """The sidecar collided with its episode and both got skipped."""
    season = tmp_path / "Show" / "Season 1"
    season.mkdir(parents=True)
    (season / "Show.S01E01.1080p.mkv").write_text("real")
    (season / "._Show.S01E01.1080p.mkv").write_text("junk")

    rows = scan.scan_tv(tmp_path)
    assert len(rows) == 1, rows
    df_tv = pd.DataFrame(rows)
    plan = excel.plan_renames(None, None, df_tv, tmp_path, NamingScheme())
    assert any(o.dst.name == "Show S01E01 [1080p].mkv" for o in plan.ops)
    assert plan.skipped == []


def test_zero_padded_season_folder_gets_normalized(tmp_path):
    show = tmp_path / "Show"
    (show / "Season 01").mkdir(parents=True)
    (show / "Season 01" / "Show.S01E01.mkv").write_text("x")

    rows = scan.scan_tv(tmp_path)
    df_tv = pd.DataFrame(rows)
    scheme = NamingScheme()
    scheme.tv_season_include_year = False
    plan = excel.plan_renames(None, None, df_tv, tmp_path, scheme)
    assert any(o.src.name == "Season 01" and o.dst.name == "Season 1"
               for o in plan.ops), [str(o.dst) for o in plan.ops]


def test_seasoning_show_is_not_a_season_folder(tmp_path):
    """_SEASONISH_DIR was an unanchored substring search."""
    show = tmp_path / "Seasoning Show"
    show.mkdir()
    (show / "Seasoning.Show.S01E01.mkv").write_text("x")
    rows = scan.scan_tv(tmp_path)
    assert [r['Show Folder'] for r in rows] == ["Seasoning Show"]
    assert rows[0]['Season'] == 1


def test_accented_titles_converge_after_one_pass(tmp_path):
    """NFD from the filesystem vs NFC from the builders re-renamed forever."""
    folder = tmp_path / unicodedata.normalize("NFD", "Amélie (2001) [1080p]")
    folder.mkdir()
    (folder / unicodedata.normalize("NFD", "Amélie (2001) [1080p].mkv")).write_text("x")

    rows = scan.scan_movies(tmp_path)
    df = pd.DataFrame(rows)
    plan = excel.plan_renames(df, tmp_path, None, None, NamingScheme())
    assert plan.ops == [], [f"{o.src} -> {o.dst}" for o in plan.ops]


def test_spreadsheet_overrides_are_sanitized(tmp_path):
    """'Folder Fixed' was taken verbatim, so a raw ':' reached the destination."""
    folder = tmp_path / "Movie.2020"
    folder.mkdir()
    (folder / "Movie.2020.mkv").write_text("x")
    rows = scan.scan_movies(tmp_path)
    df = pd.DataFrame(rows)
    df.loc[0, 'Folder Fixed'] = 'Movie: Part 2'
    plan = excel.plan_renames(df, tmp_path, None, None, NamingScheme())
    folder_ops = [o for o in plan.ops if o.src == folder]
    assert folder_ops and ':' not in folder_ops[0].dst.name
    assert folder_ops[0].dst.name == "Movie Part 2"


def test_date_based_show_is_visible_to_the_structured_scan(tmp_path):
    """These were dropped entirely, then resurfaced as S00E00."""
    show = tmp_path / "The Daily Show"
    (show / "Season 2024").mkdir(parents=True)
    (show / "Season 2024" / "The.Daily.Show.2024.01.15.Guest.mkv").write_text("x")
    rows = scan.scan_tv(tmp_path)
    assert rows and rows[0]['Episode File'], rows
    assert rows[0]['Season'] == 2024


def test_specials_folder_is_scanned_as_season_zero(tmp_path):
    show = tmp_path / "Show"
    (show / "Specials").mkdir(parents=True)
    (show / "Specials" / "Christmas Special.mkv").write_text("x")
    rows = scan.scan_tv(tmp_path)
    assert [r['Season'] for r in rows] == [0]


def test_sheet_relative_paths_use_forward_slashes(tmp_path):
    show = tmp_path / "Show"
    (show / "Season 1").mkdir(parents=True)
    (show / "Season 1" / "Show.S01E01.mkv").write_text("x")
    rows = scan.scan_tv(tmp_path)
    assert rows[0]['Episode File'] == "Season 1/Show.S01E01.mkv"
    assert "\\" not in rows[0]['Episode File']


def test_specials_are_not_given_an_invented_name(tmp_path):
    """A file with no episode number would become 'Show S00.mkv' — and two of
    them would collide and both be skipped."""
    show = tmp_path / "Show"
    (show / "Specials").mkdir(parents=True)
    (show / "Specials" / "Christmas Special.mkv").write_text("x")
    (show / "Specials" / "Behind The Scenes.mkv").write_text("x")

    rows = scan.scan_tv(tmp_path)
    assert len(rows) == 2 and all(r['Season'] == 0 for r in rows)
    plan = excel.plan_renames(None, None, pd.DataFrame(rows), tmp_path,
                              NamingScheme())
    assert plan.ops == [] and plan.skipped == []


def test_file_fixed_keeps_a_dotted_title(tmp_path):
    """Path.suffix on "Show S01E01 - Mr. Robot" is ". Robot", so a blanket
    suffix strip truncated the override to "Show S01E01 - Mr.mkv"."""
    show = tmp_path / "Show"
    (show / "Season 1").mkdir(parents=True)
    (show / "Season 1" / "Show.S01E01.mkv").write_text("x")

    rows = scan.scan_tv(tmp_path)
    df_tv = pd.DataFrame(rows)
    df_tv.loc[0, 'File Fixed'] = 'Show S01E01 - Mr. Robot'
    plan = excel.plan_renames(None, None, df_tv, tmp_path, NamingScheme())
    assert [o.dst.name for o in plan.ops] == ['Show S01E01 - Mr. Robot.mkv']


def test_file_fixed_still_strips_a_real_media_extension(tmp_path):
    show = tmp_path / "Show"
    (show / "Season 1").mkdir(parents=True)
    (show / "Season 1" / "Show.S01E01.mkv").write_text("x")

    rows = scan.scan_tv(tmp_path)
    df_tv = pd.DataFrame(rows)
    df_tv.loc[0, 'File Fixed'] = 'Renamed Episode.mkv'
    plan = excel.plan_renames(None, None, df_tv, tmp_path, NamingScheme())
    assert [o.dst.name for o in plan.ops] == ['Renamed Episode.mkv']


# --- Nested show layouts -----------------------------------------------------

def test_scan_tv_finds_a_show_behind_a_wrapper_folder(tmp_path):
    """Scanning only the top level recorded the *genre* as the show, with no
    episodes and the genre's name as the title."""
    tv = tmp_path / "TV"
    (tv / "Genre" / "Show" / "Season 1").mkdir(parents=True)
    (tv / "Genre" / "Show" / "Season 1" / "Show.S01E01.mkv").write_text("x")

    rows = scan_tv(tv)
    assert len(rows) == 1
    assert rows[0]["Show Folder"] == "Genre/Show"
    assert rows[0]["Title"] == "Show"
    assert rows[0]["Season"] == 1
    assert rows[0]["Episode File"] == "Season 1/Show.S01E01.mkv"


def test_nested_show_title_comes_from_the_show_not_the_path(tmp_path):
    """plan_renames parsed the whole relative path, so the wrapper folders
    leading to a show could end up supplying its title."""
    tv = tmp_path / "TV"
    season = tv / "Genre" / "Sub" / "Show" / "Season 1"
    season.mkdir(parents=True)
    (season / "Show.S01E01.1080p.mkv").write_text("x")

    rows = scan_tv(tv)
    plan = plan_renames(None, None, pd.DataFrame(rows), str(tv), NamingScheme())
    dsts = [op.dst.name for op in plan.ops]
    assert "Show S01E01 [1080p].mkv" in dsts


def test_nested_show_rename_stays_in_place(tmp_path):
    """A renamed nested show must keep its parent directories."""
    tv = tmp_path / "TV"
    season = tv / "Genre" / "Show.Name.2019" / "Season 1"
    season.mkdir(parents=True)
    (season / "Show.Name.S01E01.mkv").write_text("x")

    rows = scan_tv(tv)
    plan = plan_renames(None, None, pd.DataFrame(rows), str(tv), NamingScheme())
    show_ops = [op for op in plan.ops if op.src == tv / "Genre" / "Show.Name.2019"]
    assert len(show_ops) == 1
    assert show_ops[0].dst.parent == tv / "Genre"


def test_scan_tv_still_reports_folders_with_no_episodes(tmp_path):
    """Placeholder rows are what trigger the recursive fallback in run_scan."""
    tv = tmp_path / "TV"
    (tv / "Empty Show").mkdir(parents=True)
    (tv / "Real Show" / "Season 1").mkdir(parents=True)
    (tv / "Real Show" / "Season 1" / "Real.S01E01.mkv").write_text("x")

    rows = scan_tv(tv)
    by_folder = {r["Show Folder"]: r for r in rows}
    assert by_folder["Empty Show"]["Episode File"] == ""
    assert by_folder["Real Show"]["Episode File"] == "Season 1/Real.S01E01.mkv"


def test_unexplained_subtree_under_a_wrapper_still_gets_a_row(tmp_path):
    """Coverage is per subtree, not per top-level branch.

    Marking the whole of "Genre" covered as soon as one show was found there
    left any other unrecognised subtree under it with no row and no recursive
    fallback — silently invisible in the spreadsheet.
    """
    tv = tmp_path / "TV"
    (tv / "Genre" / "Real Show" / "Season 1").mkdir(parents=True)
    (tv / "Genre" / "Real Show" / "Season 1" / "RS.S01E01.mkv").write_text("x")
    (tv / "Genre" / "Odd Layout" / "disc").mkdir(parents=True)
    (tv / "Genre" / "Odd Layout" / "disc" / "whatever.mkv").write_text("x")

    rows = scan_tv(tv)
    by_folder = {r["Show Folder"]: r for r in rows}
    assert by_folder["Genre/Real Show"]["Episode File"] == "Season 1/RS.S01E01.mkv"
    assert by_folder["Genre/Odd Layout"]["Episode File"] == ""


def test_wrapper_dirs_on_the_way_to_a_show_are_not_reported_as_gaps(tmp_path):
    """"Genre" itself leads to a show, so it is explained and needs no row."""
    tv = tmp_path / "TV"
    (tv / "Genre" / "Show" / "Season 1").mkdir(parents=True)
    (tv / "Genre" / "Show" / "Season 1" / "S.S01E01.mkv").write_text("x")

    rows = scan_tv(tv)
    assert [r["Show Folder"] for r in rows] == ["Genre/Show"]


def test_scan_sees_episodes_in_a_dump_subfolder(tmp_path):
    """The scanner and the organizer must agree about what a show contains.

    scan_tv only walked top-level files and season-ish folders, so an episode
    in "Disc 1" was invisible — while the organizer would happily move it. A
    scan-then-rename run (menu [1] or [5], no organize step) skipped it, and
    the recursive fallback never fired because the show had other episodes.
    """
    tv = tmp_path / "TV"
    (tv / "Show" / "Season 1").mkdir(parents=True)
    (tv / "Show" / "Season 1" / "Show.S01E01.mkv").write_text("x")
    (tv / "Show" / "Disc 1").mkdir(parents=True)
    (tv / "Show" / "Disc 1" / "Show.S01E02.mkv").write_text("x")

    files = {r["Episode File"] for r in scan_tv(tv)}
    assert files == {"Season 1/Show.S01E01.mkv", "Disc 1/Show.S01E02.mkv"}


def test_a_nested_show_is_not_absorbed_into_its_parent(tmp_path):
    """A subfolder with season folders of its own is a show, not a dump."""
    tv = tmp_path / "TV"
    (tv / "Show").mkdir(parents=True)
    (tv / "Show" / "Show.S01E01.mkv").write_text("x")
    (tv / "Show" / "Other Show" / "Season 1").mkdir(parents=True)
    (tv / "Show" / "Other Show" / "Season 1" / "Other.S01E01.mkv").write_text("x")

    rows = {r["Show Folder"]: r["Episode File"] for r in scan_tv(tv)}
    assert rows == {"Show": "Show.S01E01.mkv",
                    "Show/Other Show": "Season 1/Other.S01E01.mkv"}


# --- Movie rename regressions ------------------------------------------------

def test_movie_renames_converge_on_a_second_pass(tmp_path):
    """Titles ending in a number were degraded by every run after the first:
    "Blade Runner 2049" lost its number, "300" grew a "() [1080p]" each time."""
    movies = tmp_path / "Movies"
    for folder, f in [("Blade.Runner.2049.2017.2160p", "Blade.Runner.2049.2017.2160p.mkv"),
                      ("300.2006.1080p", "300.2006.1080p.mkv"),
                      ("Apollo.13.1995.1080p", "Apollo.13.1995.1080p.mkv")]:
        make_movie(movies, folder, [f])

    for _ in range(2):
        rows = scan_movies(movies)
        plan = plan_renames(pd.DataFrame(rows), str(movies), None, None,
                            NamingScheme())
        for op in plan.ops:
            op.dst.parent.mkdir(parents=True, exist_ok=True)
            op.src.rename(op.dst)

    names = sorted(p.name for p in movies.rglob("*.mkv"))
    assert names == ["300 (2006) [1080p].mkv",
                     "Apollo 13 (1995) [1080p].mkv",
                     "Blade Runner 2049 (2017) [2160p].mkv"]
    # A third pass must plan nothing at all.
    rows = scan_movies(movies)
    assert plan_renames(pd.DataFrame(rows), str(movies), None, None,
                        NamingScheme()).ops == []


def test_a_comma_in_a_lone_filename_is_not_a_separator(tmp_path):
    """Splitting on the comma made two phantom sources, both skipped as
    missing, so the file was silently never renamed."""
    movies = tmp_path / "Movies"
    make_movie(movies, "Crouching Tiger Hidden Dragon (2000)",
               ["Crouching Tiger, Hidden Dragon.mkv"])
    rows = scan_movies(movies)
    plan = plan_renames(pd.DataFrame(rows), str(movies), None, None,
                        NamingScheme())
    assert plan.skipped == []
    assert any(op.dst.name == "Crouching Tiger, Hidden Dragon (2000).mkv"
               for op in plan.ops)


def test_legacy_comma_joined_list_still_splits(tmp_path):
    assert excel._video_files({'Video Files': 'A.2001.mkv, B.2002.mkv'}) == \
        ['A.2001.mkv', 'B.2002.mkv']
    assert excel._video_files({'Video Files': 'Crouching Tiger, Hidden Dragon.mkv'}) == \
        ['Crouching Tiger, Hidden Dragon.mkv']
    assert excel._video_files({'Video Files': 'A.mkv | B.mkv'}) == ['A.mkv', 'B.mkv']
    assert excel._video_files({'Video Files': ''}) == []


def test_root_level_files_do_not_plan_a_rename_of_the_library_root(tmp_path):
    """A recursive scan records root-level files under '.', which was then
    treated as a folder to rename — a move of the root into itself."""
    from mediaorg.scan import scan_recursive
    movies = tmp_path / "Movies"
    movies.mkdir()
    (movies / "Inception.2010.1080p.mkv").write_text("x")

    rows = scan_recursive(movies)
    plan = plan_renames(pd.DataFrame(rows), str(movies), None, None,
                        NamingScheme())
    assert all(op.src != movies for op in plan.ops if op.src)
    assert [op.dst.name for op in plan.ops] == ["Inception (2010) [1080p].mkv"]


def test_llm_quality_is_applied(tmp_path):
    """The prompt asks for quality and the parser returns it, but it used to
    be dropped, so an LLM-resolved name came out with no [1080p]."""
    movies = tmp_path / "Movies"
    make_movie(movies, "hopeless.name.xyz", ["hopeless.name.xyz.mkv"])
    llm_results = {
        "hopeless.name.xyz": {"title": "Hopeless Name", "year": "2011",
                              "quality": "1080p"},
        "hopeless.name.xyz.mkv": {"title": "Hopeless Name", "year": "2011",
                                  "quality": "1080p"},
    }
    rows = scan_movies(movies)
    plan = plan_renames(pd.DataFrame(rows), str(movies), None, None,
                        NamingScheme(), llm_results)
    assert sorted(op.dst.name for op in plan.ops) == \
        ["Hopeless Name (2011) [1080p]", "Hopeless Name (2011) [1080p].mkv"]
